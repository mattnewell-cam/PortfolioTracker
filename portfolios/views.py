from django.shortcuts import redirect, get_object_or_404, render
from django.http import JsonResponse
from django.contrib import messages
from django.urls import reverse_lazy
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib.auth.decorators import login_required
from django.views.decorators.http import require_POST
from django.views.generic import DetailView, CreateView, ListView
from django.db.models import Q
from django.core.serializers.json import DjangoJSONEncoder
from django.utils import timezone
from django.core.validators import validate_email
from django.core.exceptions import ValidationError
from django.conf import settings
from decimal import Decimal
from urllib.parse import urlparse, urlunparse
import random
import feedparser

from core.yfinance_client import get_quote
from core.email import send_email
from .models import Portfolio, Order, PortfolioSnapshot, PortfolioFollower, PortfolioAllowedEmail, NotificationSetting
from .constants import BENCHMARK_CHOICES
from .forms import (
    PortfolioForm,
    OrderForm,
    AllowedEmailForm,
    AllowedEmailUploadForm,
    AccountForm,
    NotificationSettingForm,
)
from core.forms import EmailVerificationForm
import yfinance as yf
import json
import csv
from io import StringIO

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None


def _fetch_substack_metadata(substack_url):
    parsed_url = urlparse(substack_url)
    feed_url = urlunparse(
        parsed_url._replace(path="/feed", params="", query="", fragment="")
    )
    title = None
    subtitle = None
    try:
        feed = feedparser.parse(feed_url)
        title = feed.feed.get("title")
        subtitle = feed.feed.get("subtitle") or feed.feed.get("description")
    except Exception:
        pass
    return title, subtitle


def _get_position_value(symbol, qty):
    """Return tuple of (mid_local, currency, fx_rate, value_usd) for a holding."""
    try:
        quote = get_quote(symbol)
        price_val = quote.get("price")
        currency = quote.get("currency")
        fx_rate_val = quote.get("fx_rate")
        if price_val is not None and fx_rate_val is not None:
            mid_local = Decimal(str(price_val))
            fx_rate = Decimal(str(fx_rate_val))
            value_usd = mid_local * fx_rate * Decimal(str(qty))
            return mid_local, currency, fx_rate, value_usd
    except Exception:
        pass
    return None, None, None, None


def build_portfolio_context(p, include_details=True):
    """Return context data for a portfolio."""
    positions = []
    total_value = p.cash_balance
    for symbol, qty in p.holdings.items():
        mid_local = currency = fx_rate = value_usd = None
        if include_details:
            mid_local, currency, fx_rate, value_usd = _get_position_value(symbol, qty)
        else:
            _, _, _, value_usd = _get_position_value(symbol, qty)

        if include_details:
            positions.append({
                "symbol": symbol,
                "quantity": qty,
                "mid_local": mid_local,
                "currency": currency,
                "fx_rate": fx_rate,
                "value_usd": value_usd,
            })

        if value_usd is not None:
            total_value += value_usd

    cash_allocation = None
    if include_details and total_value > 0:
        for pos in positions:
            if pos["value_usd"] is not None:
                pos["allocation"] = (pos["value_usd"] / total_value) * 100
            else:
                pos["allocation"] = None
        cash_allocation = (p.cash_balance / total_value) * 100

    orders_data = []
    if include_details:
        for o in p.orders.all().order_by("-executed_at"):
            total_usd = o.price_executed * o.fx_rate * o.quantity
            orders_data.append({
                "executed_at": o.executed_at,
                "symbol": o.symbol,
                "side": o.side,
                "quantity": o.quantity,
                "price_local": o.price_executed,
                "currency": o.currency,
                "fx_rate": o.fx_rate,
                "total_value_usd": total_usd,
            })

    history_data = []
    for snap in p.snapshots.all().order_by("timestamp"):
        history_data.append({
            "date": snap.timestamp.date().isoformat(),
            "value": snap.total_value,
        })
    if not history_data:
        history_data.append({
            "date": timezone.now().date().isoformat(),
            "value": total_value,
        })

    benchmark_data = []
    snaps = list(p.snapshots.all().order_by("timestamp"))
    if snaps:
        per_ticker = {t: [] for t, _ in BENCHMARK_CHOICES}
        for snap in snaps:
            date_iso = snap.timestamp.date().isoformat()
            for ticker, _ in BENCHMARK_CHOICES:
                price = snap.benchmark_values.get(ticker)
                if price is not None:
                    per_ticker[ticker].append({"date": date_iso, "price_usd": price})
        for ticker, name in BENCHMARK_CHOICES:
            pts = per_ticker[ticker]
            if pts:
                base = pts[0]["price_usd"]
                for pt in pts:
                    pt["price_usd"] = (pt["price_usd"] / base) * 100_000
            benchmark_data.append({
                "ticker": ticker,
                "label": name,
                "data": pts,
            })
    else:
        for ticker, name in BENCHMARK_CHOICES:
            benchmark_data.append({
                "ticker": ticker,
                "label": name,
                "data": [],
            })

    default_benchmarks = p.benchmarks

    if benchmark_data and default_benchmarks:
        default_set = set(default_benchmarks)
        ordered_data = []

        for ticker, _ in BENCHMARK_CHOICES:
            if ticker in default_set:
                match = next((bm for bm in benchmark_data if bm["ticker"] == ticker), None)
                if match:
                    ordered_data.append(match)

        for bm in benchmark_data:
            if bm["ticker"] not in default_set:
                ordered_data.append(bm)

        benchmark_data = ordered_data

    return {
        "positions": positions if include_details else [],
        "total_value": total_value,
        "cash_allocation": cash_allocation if include_details else None,
        "orders_data": orders_data if include_details else [],
        "history_data": history_data,
        "history_data_json": json.dumps(history_data, cls=DjangoJSONEncoder),
        "benchmark_data": benchmark_data,
        "benchmark_data_json": json.dumps(benchmark_data, cls=DjangoJSONEncoder),
        "default_benchmarks": default_benchmarks,
    }


class PortfolioCreateView(LoginRequiredMixin, CreateView):
    model = Portfolio
    form_class = PortfolioForm
    template_name = "portfolios/portfolio_form.html"
    success_url = reverse_lazy("portfolios:portfolio-detail")

    def dispatch(self, request, *args, **kwargs):
        if Portfolio.objects.filter(user=request.user).exists():
            return redirect("portfolios:portfolio-detail")
        return super().dispatch(request, *args, **kwargs)

    def form_valid(self, form):
        form.instance.user = self.request.user
        return super().form_valid(form)


class PortfolioDetailView(LoginRequiredMixin, DetailView):
    model = Portfolio
    template_name = "portfolios/portfolio_detail.html"
    context_object_name = "portfolio"

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_authenticated and not Portfolio.objects.filter(
            user=request.user, is_deleted=False
        ).exists():
            return render(request, "portfolios/portfolio_empty.html")
        return super().dispatch(request, *args, **kwargs)

    def get_object(self, queryset=None):
        return get_object_or_404(Portfolio, user=self.request.user, is_deleted=False)

    
    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx.update(build_portfolio_context(self.object))
        ctx["is_owner"] = True
        ctx["private_view"] = False
        ctx["allowed_count"] = self.object.allowed_emails.count()
        ctx["followers_count"] = self.object.followers.count()
        ctx["order_form"] = OrderForm()
        return ctx


class PublicPortfolioDetailView(DetailView):
    model = Portfolio
    template_name = "portfolios/portfolio_detail.html"
    context_object_name = "portfolio"
    slug_field = "url_tag"
    slug_url_kwarg = "tag"

    def get_queryset(self):
        return Portfolio.objects.filter(is_deleted=False)

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        is_owner = (
            self.request.user.is_authenticated and self.request.user == self.object.user
        )
        is_allowed = False
        if (
            self.object.is_private
            and self.request.user.is_authenticated
            and self.request.user != self.object.user
        ):
            identifier = self.request.user.email or self.request.user.username
            is_allowed = self.object.allowed_emails.filter(
                email=identifier
            ).exists()
        include_details = is_owner or not self.object.is_private or is_allowed
        ctx.update(build_portfolio_context(self.object, include_details=include_details))
        ctx["is_owner"] = is_owner
        ctx["is_allowed"] = is_allowed
        ctx["private_view"] = self.object.is_private and not include_details
        if self.request.user.is_authenticated:
            ctx["is_following"] = self.object.followers.filter(
                follower=self.request.user
            ).exists()
        else:
            ctx["is_following"] = False
        ctx["allowed_count"] = self.object.allowed_emails.count()
        ctx["followers_count"] = self.object.followers.count()
        return ctx


@require_POST
@login_required
def toggle_privacy(request):
    """Toggle the current user's portfolio privacy flag."""
    portfolio = get_object_or_404(
        Portfolio, user=request.user, is_deleted=False
    )
    choice = request.POST.get("privacy_choice")

    if portfolio.is_private:
        portfolio.is_private = False
        portfolio.save(update_fields=["is_private"])
        return redirect("portfolios:portfolio-detail")

    if choice == "allow_followers":
        for follower_rel in portfolio.followers.select_related("follower"):
            follower = follower_rel.follower
            identifier = follower.email or follower.username
            if identifier:
                PortfolioAllowedEmail.objects.get_or_create(
                    portfolio=portfolio, email=identifier
                )
    elif choice == "remove_followers":
        follower_identifiers = []
        for follower_rel in portfolio.followers.select_related("follower"):
            follower = follower_rel.follower
            identifier = follower.email or follower.username
            if identifier:
                follower_identifiers.append(identifier)
        portfolio.followers.all().delete()
        if follower_identifiers:
            portfolio.allowed_emails.filter(email__in=follower_identifiers).delete()

    portfolio.is_private = True
    portfolio.save(update_fields=["is_private"])
    return redirect("portfolios:portfolio-detail")


@require_POST
@login_required
def toggle_follow(request, tag):
    portfolio = get_object_or_404(Portfolio, url_tag=tag, is_deleted=False)
    if portfolio.user == request.user:
        return redirect("portfolios:portfolio-public-detail", tag=tag)
    identifier = request.user.email or request.user.username
    allowed = (
        not portfolio.is_private
        or portfolio.allowed_emails.filter(email=identifier).exists()
    )
    if not allowed:
        return redirect("portfolios:portfolio-public-detail", tag=tag)
    rel, created = PortfolioFollower.objects.get_or_create(
        portfolio=portfolio, follower=request.user
    )
    if not created:
        rel.delete()
    return redirect("portfolios:portfolio-public-detail", tag=tag)


@login_required
def lookup_quote(request):
    symbol = request.GET.get("symbol", "").strip()
    if not symbol:
        return JsonResponse({"error": "Please enter a ticker symbol."}, status=400)

    try:
        quote = get_quote(symbol)
    except Exception:
        return JsonResponse({"error": "Unable to fetch quote for that ticker."}, status=400)

    display_name = quote.get("longName") or quote.get("shortName")
    price = quote.get("price")
    currency = quote.get("currency")
    fx_rate = quote.get("fx_rate")

    if display_name is None or price is None or currency is None or fx_rate is None:
        return JsonResponse({"error": "Ticker not found. Please try another."}, status=404)

    market_state = quote.get("market_state")
    return JsonResponse(
        {
            "symbol": (quote.get("symbol") or symbol).upper(),
            "name": display_name,
            "price": price,
            "currency": currency,
            "fx_rate": fx_rate,
            "market_state": market_state,
            "market_open": market_state == "REGULAR",
        }
    )


@login_required
def allow_list(request):
    portfolio = get_object_or_404(
        Portfolio, user=request.user, is_private=True, is_deleted=False
    )
    emails = portfolio.allowed_emails.all().order_by("email")
    email_form = AllowedEmailForm()
    upload_form = AllowedEmailUploadForm()
    if request.method == "POST":
        action = request.POST.get("action")
        if action == "add_email":
            email_form = AllowedEmailForm(request.POST)
            if email_form.is_valid():
                PortfolioAllowedEmail.objects.get_or_create(
                    portfolio=portfolio, email=email_form.cleaned_data["email"]
                )
                return redirect("portfolios:portfolio-allow-list")
        elif action == "upload":
            upload_form = AllowedEmailUploadForm(request.POST, request.FILES)
            if upload_form.is_valid():
                file = upload_form.cleaned_data["file"]
                try:
                    emails_raw = []
                    file_name = file.name.lower()
                    if file_name.endswith(".csv") or file_name.endswith(".tsv"):
                        file.seek(0)
                        data = file.read().decode("utf-8")
                        delimiter = "\t" if file_name.endswith(".tsv") else ","
                        reader = csv.reader(StringIO(data), delimiter=delimiter)
                        emails_raw = [row[0] for row in reader if row]
                    else:
                        if load_workbook is None:
                            raise ImportError("openpyxl is required for Excel uploads")
                        file.seek(0)
                        wb = load_workbook(file)
                        ws = wb.active
                        emails_raw = [
                            row[0]
                            for row in ws.iter_rows(min_col=1, max_col=1, values_only=True)
                            if row
                        ]
                    for e in emails_raw:
                        if not e:
                            continue
                        email = str(e).strip().lstrip("\ufeff")
                        try:
                            validate_email(email)
                        except ValidationError:
                            continue
                        PortfolioAllowedEmail.objects.get_or_create(
                            portfolio=portfolio, email=email
                        )
                except Exception:
                    pass
                return redirect("portfolios:portfolio-allow-list")
        elif action == "delete":
            email_id = request.POST.get("id")
            PortfolioAllowedEmail.objects.filter(
                id=email_id, portfolio=portfolio
            ).delete()
            return redirect("portfolios:portfolio-allow-list")
        elif action == "delete_all":
            portfolio.allowed_emails.all().delete()
            return redirect("portfolios:portfolio-allow-list")
    return render(
        request,
        "portfolios/allow_list.html",
        {
            "portfolio": portfolio,
            "emails": emails,
            "email_form": email_form,
            "upload_form": upload_form,
        },
    )


class PortfolioExploreView(ListView):
    model = Portfolio
    template_name = "portfolios/portfolio_explore.html"
    context_object_name = "portfolios"

    def get_queryset(self):
        qs = Portfolio.objects.filter(is_deleted=False)
        query = self.request.GET.get("q")
        if query:
            qs = qs.filter(Q(name__icontains=query) | Q(substack_url__icontains=query))
        return qs.order_by("-created_at")

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        for p in ctx["portfolios"]:
            snap = p.snapshots.order_by("-timestamp").first()
            if snap:
                p.total_value_cached = snap.total_value
            else:
                p.total_value_cached = build_portfolio_context(
                    p, include_details=False
                )["total_value"]
        ctx["search_query"] = self.request.GET.get("q", "")
        return ctx


class FollowedPortfoliosView(LoginRequiredMixin, ListView):
    model = Portfolio
    template_name = "portfolios/portfolio_followed.html"
    context_object_name = "portfolios"

    def get_queryset(self):
        return (
            Portfolio.objects.filter(
                followers__follower=self.request.user, is_deleted=False
            )
            .order_by("-created_at")
        )

    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        for p in ctx["portfolios"]:
            snap = p.snapshots.order_by("-timestamp").first()
            if snap:
                p.total_value_cached = snap.total_value
            else:
                p.total_value_cached = p.cash_balance
        return ctx


class OrderCreateView(LoginRequiredMixin, CreateView):
    model = Order
    form_class = OrderForm
    template_name = "portfolios/portfolio_detail.html"

    def dispatch(self, request, *args, **kwargs):
        self.portfolio = get_object_or_404(
            Portfolio, user=request.user, is_deleted=False
        )
        if request.method.lower() == "get":
            return redirect("portfolios:portfolio-detail")
        return super().dispatch(request, *args, **kwargs)


    def get_context_data(self, **kwargs):
        ctx = super().get_context_data(**kwargs)
        ctx["portfolio"] = self.portfolio
        ctx.update(build_portfolio_context(self.portfolio))
        ctx["is_owner"] = True
        ctx["private_view"] = False
        ctx["allowed_count"] = self.portfolio.allowed_emails.count()
        ctx["order_form"] = kwargs.get("form", OrderForm())
        return ctx

    #-------------

    def form_valid(self, form):
        print("Running form_valid")
        form.instance.portfolio = self.portfolio

        symbol = form.cleaned_data["symbol"].upper()
        side = form.cleaned_data["side"]  # "BUY" or "SELL"
        quantity = form.cleaned_data["quantity"]

        print(symbol, side, quantity)

        # 1) Fetch price
        try:
            quote = get_quote(symbol)
            price = Decimal(str(quote["price"]))
            bid = Decimal(str(quote["bid"])) if quote.get("bid") is not None else None
            ask = Decimal(str(quote["ask"])) if quote.get("ask") is not None else None
            traded_today = quote["traded_today"]
            currency = quote["currency"]
            fx_rate = Decimal(str(quote["fx_rate"]))
            market_state = quote["market_state"]
            print(price, bid, ask, traded_today, currency, fx_rate)
        except Exception:
            form.add_error(None, f"Could not fetch live quote for “{symbol}”.")
            return self.form_invalid(form)

        if not settings.DEBUG and market_state != "REGULAR":
            form.add_error(
                None,
                "Order failed because the market is currently closed."
            )
            return self.form_invalid(form)

        # 2) Compute execution_price & validate
        if side == "BUY":
            execution_price = price if traded_today else ask
            total_cost = execution_price * fx_rate * quantity
            if self.portfolio.cash_balance < total_cost:
                form.add_error(
                    None,
                    f"Insufficient cash: need ${total_cost:.2f}, have ${self.portfolio.cash_balance:.2f}."
                )
                return self.form_invalid(form)

        else:  # SELL
            execution_price = price if traded_today else bid
            held_qty = self.portfolio.holdings.get(symbol, 0)
            if held_qty < quantity:
                form.add_error(
                    None,
                    f"Cannot sell {quantity} shares of {symbol}; you only hold {held_qty}."
                )
                return self.form_invalid(form)

        # 3) Assign price and save Order
        form.instance.price_executed = execution_price
        form.instance.fx_rate = fx_rate
        form.instance.currency = currency
        response = super().form_valid(form)

        # 4) Update cash_balance and holdings
        if side == "BUY":
            self.portfolio.cash_balance -= execution_price * fx_rate * quantity
            new_qty = float(self.portfolio.holdings.get(symbol, 0)) + quantity
            self.portfolio.holdings[symbol] = new_qty
        else:  # SELL
            self.portfolio.cash_balance += execution_price * fx_rate * quantity
            remaining = float(self.portfolio.holdings.get(symbol, 0)) - quantity
            if remaining == 0:
                # remove key if zero
                self.portfolio.holdings.pop(symbol, None)
            else:
                self.portfolio.holdings[symbol] = remaining

        self.portfolio.save()
        follower_emails = []
        for follower_rel in self.portfolio.followers.select_related(
            "follower__notification_setting"
        ):
            follower = follower_rel.follower
            setting = getattr(follower, "notification_setting", None)
            preference = (
                setting.preference
                if setting
                else NotificationSetting.PREFERENCE_IMMEDIATE
            )
            if preference == NotificationSetting.PREFERENCE_IMMEDIATE and follower.email:
                follower_emails.append(follower.email)

        verb = "bought" if side == "BUY" else "sold"
        if follower_emails:
            send_email(
                "notifications@trackstack.uk",
                f"New trade in {self.portfolio.name}'s Portfolio",
                f"{self.portfolio.name} {verb} {quantity} shares of {symbol} at {currency} {round(execution_price, 3)}\n\n ",
                follower_emails,
                fail_silently=True,
            )
        return response



    def form_invalid(self, form):
        """
        If validation fails above, form.add_error(…) was called.
        Returning form_invalid(form) will re-render the page with errors.
        """
        return super().form_invalid(form)

    def get_success_url(self):
        return reverse_lazy("portfolios:portfolio-detail")


@login_required
def portfolio_history(request):
    p = Portfolio.objects.filter(user=request.user, is_deleted=False).first()
    if not p:
        return JsonResponse({"error": "Not found"}, status=404)

    data = [
        {
            "timestamp": snap.timestamp.isoformat(),
            "value": snap.total_value,
        }
        for snap in p.snapshots.all()
    ]

    # Create single datapoint if no snapshots
    if not data:
        total_value = p.cash_balance
        for symbol, qty in p.holdings.items():
            try:
                quote = get_quote(symbol)
                total_value += Decimal(str(quote["price"])) * Decimal(str(quote["fx_rate"])) * Decimal(str(qty))
            except Exception:
                pass
        data.append({
            "timestamp": timezone.now().isoformat(),
            "value": total_value,
        })


    return JsonResponse(data, safe=False)


@login_required
def account_details(request):
    portfolio = Portfolio.objects.filter(user=request.user).first()
    notification_form = NotificationSettingForm(user=request.user)
    form = AccountForm(instance=request.user)
    if request.method == "POST":
        action = request.POST.get("action")
        if action == "notifications":
            notification_form = NotificationSettingForm(
                request.POST, user=request.user
            )
            if notification_form.is_valid():
                notification_form.save()
                messages.success(
                    request, "Notification preferences updated successfully."
                )
                return redirect("portfolios:account-details")
        elif action == "delete_portfolio":
            if portfolio and not portfolio.is_deleted:
                portfolio.is_deleted = True
                portfolio.is_private = True
                portfolio.deleted_at = timezone.now()
                portfolio.save(update_fields=["is_deleted", "is_private", "deleted_at"])
                messages.success(
                    request,
                    "Your portfolio has been deleted and removed from public pages. "
                    "Its data is retained in case you need to restore it later.",
                )
            return redirect("portfolios:account-details")
        elif action == "refresh_substack_name":
            if portfolio and portfolio.substack_url:
                title, subtitle = _fetch_substack_metadata(portfolio.substack_url)
                if title:
                    update_fields = []
                    if title != portfolio.name:
                        portfolio.name = title
                        update_fields.append("name")
                    if subtitle is not None and subtitle != portfolio.short_description:
                        portfolio.short_description = subtitle
                        update_fields.append("short_description")
                    if update_fields:
                        portfolio.save(update_fields=update_fields)
                    messages.success(request, "Substack name updated.")
                else:
                    messages.error(
                        request, "Unable to update Substack name. Please try again later."
                    )
            else:
                messages.error(request, "No portfolio found to update.")
            return redirect("portfolios:account-details")
        else:
            form = AccountForm(request.POST, instance=request.user)
            old_email = request.user.email
            if form.is_valid():
                user = request.user
                new_display = form.cleaned_data.get("display_name", "")
                new_email = form.cleaned_data.get("email") or old_email
                email_changed = new_email.lower() != old_email.lower()
                if new_display != user.first_name:
                    user.first_name = new_display
                    user.save(update_fields=["first_name"])
                if email_changed:
                    code = f"{random.randint(0, 999999):06d}"
                    request.session["pending_email_change"] = {
                        "new_email": new_email,
                        "code": code,
                    }
                    send_email(
                        "verify@trackstack.uk",
                        "Verify your new email",
                        f"Your verification code is {code}",
                        [new_email],
                        fail_silently=True,
                    )
                    return redirect("portfolios:account-verify-email")
                return redirect("portfolios:account-details")
    else:
        form = AccountForm(instance=request.user)
    return render(
        request,
        "portfolios/account_details.html",
        {"portfolio": portfolio, "form": form, "notification_form": notification_form},
    )


@login_required
def verify_email_change(request):
    pending = request.session.get("pending_email_change")
    if not pending:
        return redirect("portfolios:account-details")

    if request.method == "POST" and request.POST.get("resend"):
        code = f"{random.randint(0, 999999):06d}"
        pending["code"] = code
        request.session["pending_email_change"] = pending
        send_email(
            "verify@trackstack.uk",
            "Verify your new email",
            f"Your verification code is {code}",
            [pending["new_email"]],
            fail_silently=True,
        )
        messages.success(request, "A new verification code has been sent.")
        return redirect("portfolios:account-verify-email")

    if request.method == "POST":
        form = EmailVerificationForm(request.POST)
        if form.is_valid():
            if form.cleaned_data["code"] == pending["code"]:
                user = request.user
                user.email = pending["new_email"]
                user.username = pending["new_email"]
                user.save()
                del request.session["pending_email_change"]
                return redirect("portfolios:account-details")
            else:
                form.add_error("code", "Invalid verification code.")
    else:
        form = EmailVerificationForm()
    return render(
        request,
        "portfolios/verify_email_change.html",
        {"form": form, "new_email": pending["new_email"]},
    )
